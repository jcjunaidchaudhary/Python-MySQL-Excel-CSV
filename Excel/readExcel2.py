import openpyxl

wb_obj=openpyxl.load_workbook("School.xlsx")
sheet_obj=wb_obj.active

# cell_obj=sheet_obj.cell(row=1,column=1)
# print(cell_obj.value)

# for i in range(1,sheet_obj.max_column+1):
#     cell_obj=sheet_obj.cell(row=1,column=i)
#     print(cell_obj.value,end=" ")
lst=[]
for i in range(2,sheet_obj.max_row):
    for j in range(1,sheet_obj.max_column+1):
        cell_obj=sheet_obj.cell(row=i,column=j)
    #     print(cell_obj.value,end="   ")
    # print("\n")
        lst.append(cell_obj.value)
print(lst)