
# Python program to read an excel file

# import openpyxl module
import openpyxl
from openpyxl.worksheet import worksheet

# Give the location of the file
path = "School3.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# dict_data=[]
coloumn=['name','Roll_no','class','percentage','contact']

def addingDB(param): 
    print("param", param)
       
    # zip_data=dict(zip(coloumn,param))
    # dict_data.append(zip_data)
    row_value.clear()

line=0
row_value=[]
for row in sheet_obj.values:
    print(row)
    line += 1
    for value in row:
       if line==1:
           continue
       else:
           row_value.append(value)
    if line != 1:
        addingDB(row_value)

print("row_value", row_value)
# print("dict_data", dict_data)


# for row in sheet_obj.iter_rows(values_only=True):
#     lst_row=list(row)
#     addingDB(lst_row,dic_data)

# print(dic_data)


# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method.
# cell_obj = sheet_obj.cell(row = 1, column = 1)

# Print value of cell object
# using the value attribute
# print(cell_obj.value)
# print(sheet_obj.max_row)
# print(sheet_obj.max_column)
# c=worksheet['name']
# print(c)
