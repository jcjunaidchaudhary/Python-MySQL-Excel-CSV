import mysql.connector,openpyxl
from openpyxl.workbook import workbook

conn=mysql.connector.connect(
    host="localhost",
    user="root",
    passwd="root",
    database="School"
)
cursor=conn.cursor()
wb_obj=openpyxl.load_workbook("School3.xlsx")
sheet=wb_obj.active

cursor.execute("select Roll_no from SchoolTb")
myresult=cursor.fetchall()
existing=[]
for x in myresult:
    lst=list(x)
    existing+=lst
print("existing", existing)

updatequery="""update SchoolTb 
        set name=%s, class=%s, percentage=%s, contact=%s
        where Roll_no=%s"""
insertquery="""INSERT INTO SchoolTb (name, Roll_no, class, percentage, contact) 
        VALUES (%s,%s,%s,%s,%s)"""
                               
row=0

#code
for r in range(2,sheet.max_row+1):
    name    =sheet.cell(r,1).value
    roll =sheet.cell(r,2).value
    Class =sheet.cell(r,3).value
    percentage=sheet.cell(r,4).value
    contact =sheet.cell(r,5).value
    
    insertvalues=(name,roll,Class,percentage,contact)
    updatevalues=(name,Class,percentage,contact,roll)

    if roll in existing:
        cursor.execute(updatequery,updatevalues)
        row+=cursor.rowcount     
    else:        
        cursor.execute(insertquery,insertvalues)
        row+=cursor.rowcount


cursor.close()
conn.commit()
conn.close()
print(row, "was inserted.")


# query="""INSERT IGNORE INTO SchoolTb (name, Roll_no, class, percentage, contact)
#      VALUES (%s,%s,%s,%s,%s)"""
# query="""REPLACE INTO SchoolTb (name, Roll_no, class, percentage, contact)
#      VALUES (%s,%s,%s,%s,%s)"""
# row=0
# # query="""SET name=%s, Roll_no=%s,class=%s, percentage=%s, contact=%s 
# #         INSERT INTO SchoolTb(name, Roll_no, class, percentage, contact)
# #      VALUES (%s,%s,%s,%s,%s) 
# #      ON DUPLICATE KEY UPDATE name=%s, class=%s, percentage=%s, contact=%s;"""
# for r in range(2,sheet.max_row+1):
#     name    =sheet.cell(r,1).value
#     roll =sheet.cell(r,2).value
#     Class =sheet.cell(r,3).value
#     percentage=sheet.cell(r,4).value
#     contact =sheet.cell(r,5).value
#     values=(name,roll,Class,percentage,contact)
#     cursor.execute(query,values)
#     row+=cursor.rowcount
